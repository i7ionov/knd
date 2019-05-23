# -*- coding: utf-8 -*-
import os
import uuid
import datetime

from django.conf import settings
from django.apps import apps
from django.db.models import Q

import iggn_tools.tools
from dictionaries.tools import normalize_house_number
from iggn_tools import filter
import xlrd, openpyxl
from openpyxl.styles import Alignment, NamedStyle
import django
from django.core.files.base import ContentFile
from openpyxl.writer.excel import save_virtual_workbook

import inspections.models
import dictionaries.models
import analytic.models
from django.core.exceptions import MultipleObjectsReturned
from datetime import datetime
from django.utils import timezone


def import_insp_from_gis_gkh(file):
    rb = xlrd.open_workbook(file)
    sheet = rb.sheet_by_index(0)
    row = 1
    while row + 1 < sheet.nrows - 1 and sheet.row_values(row + 1)[0] != '':
        row = row + 1
        val = sheet.row_values(row)
        if val[4] == '':
            print("Пропущена запись №" + str(val[0]))
            row = row + 1
            continue
        number = val[4].replace('Распоряжение № ', '').replace(' ', '')
        year = datetime.strptime(val[5], '%d.%m.%Y').year
        insp = inspections.models.Inspection.objects.filter(doc_date__year=year, doc_number__iexact=number)
        # здесь мы пытаемся найти именно ту проверку, которую нужно обновить, т.к. этот фильтр может возвратить
        # несколько проверок
        # самый гарантированный способ - номер ГИС ЖКХ, но если он еще не обновлен - пытаемся сопоставить по организации
        # сначала ищем по огрн, если не находим, то по наименованию, если не находим то последнюю без организации
        try:
            insp = insp.get(gis_gkh_number=val[1])
        except inspections.models.Inspection.DoesNotExist:
            insp2 = insp.filter(organization__ogrn=val[13]).last()
            if insp2:
                insp = insp2
            else:
                insp2 = insp.filter(organization__name__icontains=val[12]).last()
                if insp2:
                    insp = insp2
                else:
                    insp = insp.filter(organization=None).last()

        if insp is None:
            continue
            insp = inspections.models.Inspection()
            insp.doc_type = 'проверка'
            insp.doc_number = number
            insp.doc_date = datetime.strptime(val[5], '%d.%m.%Y').date()
        insp.gis_gkh_number = val[1]
        insp.erp_number = val[3]
        insp.doc_type = "проверка"
        insp.control_form = inspections.models.ControlForm.objects.get(text=val[8])
        insp.control_plan = inspections.models.ControlPlan.objects.get(text=val[6])
        if val[7] == "Государственный и муниципальный жилищный надзор (контроль)":
            insp.control_kind = inspections.models.ControlKind.objects.get(pk=1)
        elif val[7] == "Лицензионный контроль":
            insp.control_kind = inspections.models.ControlKind.objects.get(pk=2)
        # если не введен ОГРН, то это, скорее всего гражданин
        # или если ОГРН введен, но у нас такого нет
        if val[13] != "" and dictionaries.models.Organization.objects.filter(ogrn=val[13]).count() == 0:
            # ищем или создаем по названию
            try:
                org, created = dictionaries.models.Organization.objects.get_or_create(name=val[12])
                if created:
                    org.name = val[12]
                    org.inn = "ОГРН: " + val[13]
                    org.ogrn = val[13]
                    org.kpp = val[14]
                    if val[11] == 'Организация':
                        org.org_type_id = 2
                    else:
                        org.org_type_id = 3
                    org.save()
                insp.organization = org
            except MultipleObjectsReturned:
                print("Несколько организаций с именем " + val[12])
        # если у нас эта организация уже есть
        elif dictionaries.models.Organization.objects.filter(ogrn=val[13]).count() == 1 and val[11] != 'Гражданин':
            # находим ее по огрн
            org = dictionaries.models.Organization.objects.get(ogrn=val[13])
            insp.organization = org
        # если это гражданин
        else:
            # ищем или создаем по названию
            try:
                org, created = dictionaries.models.Organization.objects.get_or_create(name=val[12])
                if created:
                    org.name = val[12]
                    org.inn = val[12]
                    org.ogrn = val[12]
                    org.org_type_id = 1
                    org.save()
                insp.organization = org
            except MultipleObjectsReturned:
                print("Несколько организаций с именем " + val[12])

        try:
            if val[16] != '':
                insp.legal_basis, created = inspections.models.LegalBasis.objects.get_or_create(text=val[16])
        except:
            print("Не удалость сохранить основание для №" + str(val[0]))

        if val[17] != '':
            insp.date_begin = datetime.strptime(val[17], '%d.%m.%Y').date()
        if val[18] != '':
            insp.date_end = datetime.strptime(val[18], '%d.%m.%Y').date()
        if val[28] != '':
            insp.act_date = datetime.strptime(val[28], '%d.%m.%Y').date()
            if insp.inspection_result is None:
                insp.inspection_result_id = 1
        insp.save()


def import_addr_from_gis_gkh(file):
    rb = xlrd.open_workbook(file)
    sheet = rb.sheet_by_index(1)
    row = 2
    while row < sheet.nrows and sheet.row_values(row)[0] != '':
        val = sheet.row_values(row)
        if val[2] == '':
            print("Пропущена запись №" + str(val[0]))
            row = row + 1
            continue
        street = val[7].replace('пр-кт.', 'пр-кт') \
            .replace('б-р.', 'б-р') \
            .replace('проезд.', 'проезд') \
            .replace('тракт.', 'тракт') \
            .replace('ул. Барамзиной', 'ул. Барамзиной Татьяны') \
            .replace('ул. Алексея Кирьянова', 'ул. Кирьянова') \
            .replace('наб. Ириловская', 'Ириловская набережная') \
            .replace('наб. Иренская', 'Иренская набережная') \
            .replace('тракт Ленский', 'Ленский тракт') \
            .replace('л. Им газ. Правда', 'ул. им. газ. Правда')
        if val[5] != '':
            try:
                city = val[5].replace('п. ', '')
                addr = dictionaries.models.Address.objects.get(city=city, street=street)
                house, created = dictionaries.models.House.objects.get_or_create(address=addr, number=val[8])
                insp = inspections.models.Inspection.objects.get(gis_gkh_number=val[1])
                insp.houses.add(house)
                insp.save()
            except dictionaries.models.Address.DoesNotExist:
                print("Не найдена запись: " + val[5] + val[7])
            except inspections.models.Inspection.DoesNotExist:
                print("Не найдена проверка: " + val[1])
            except MultipleObjectsReturned:
                print("Несколько адресов " + val[5] + val[7])
        else:
            try:
                city = val[6].replace('п. ', '') \
                    .replace('пгт. Сарс', 'р.п. Сарс') \
                    .replace('пгт. ', '') \
                    .replace('рп.', '')
                area = val[4].replace('р-н. ', '')
                addr = dictionaries.models.Address.objects.get(area__contains=area, city__contains=city, street=street)
                house, created = dictionaries.models.House.objects.get_or_create(address=addr, number=val[8])
                insp = inspections.models.Inspection.objects.get(gis_gkh_number=val[1])
                insp.houses.add(house)
                insp.save()
            except dictionaries.models.Address.DoesNotExist:
                print("Не найдена запись: " + area + city + street)
            except inspections.models.Inspection.DoesNotExist:
                print("Не найдена проверка: " + val[1])
            except MultipleObjectsReturned:
                print("Несколько адресов " + area + city + street)

        row = row + 1


def import_order_from_gis_gkh(file):
    rb = xlrd.open_workbook(file)
    sheet = rb.sheet_by_index(2)
    row = 1
    while row + 1 < sheet.nrows - 1 and sheet.row_values(row + 1)[0] != '':
        row = row + 1
        val = sheet.row_values(row)
        try:
            number = val[2].lower().replace(' ', '')
            year = datetime.strptime(val[3], '%d.%m.%Y').year
            precept = inspections.models.Precept.objects.filter(doc_date__year=year, doc_number__iexact=number).last()
            if precept is None:
                continue
                precept = inspections.models.Precept()
                precept.doc_type = 'предписание'
                precept.doc_number = number
                precept.doc_date = datetime.strptime(val[3], '%d.%m.%Y').date()
                insp = inspections.models.Inspection.objects.get(gis_gkh_number=val[1])
                precept.organization = insp.organization
                precept.parent = insp
                precept.save()
                precept.houses.set(precept.parent.inspection.houses.all())

            if precept.precept_begin_date is None:
                precept.precept_begin_date = precept.doc_date
            if val[6] != '' and precept.precept_end_date is None:
                precept.precept_end_date = datetime.strptime(val[6], '%d.%m.%Y').date()
            if val[7] != '' and (precept.precept_result_id == 1 or precept.precept_result is None):
                precept.precept_result_id = 2
            precept.save()
        except inspections.models.Inspection.DoesNotExist:
            print("Not found inspection: " + val[1])
        except MultipleObjectsReturned:
            pass


def export_excel(query_set, user_id, request_post=None):
    result = analytic.models.ExportResult()
    result.user = dictionaries.models.User.objects.get(django_user__pk=user_id)
    count = query_set.count()
    if query_set.model == inspections.models.Inspection:
        wb = openpyxl.load_workbook(os.path.join(settings.MEDIA_ROOT, 'templates', 'inspections.xlsx'))
    else:
        wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "iggndb"
    fields = filter.get_model_columns([], query_set.model)
    # TODO: убрать костыль
    if query_set.model == inspections.models.Inspection:
        fields.append(
            {'verbose_name': 'номер предписания', 'name': 'doc_number', 'prefix': 'children.', 'field': 'custom'})
    if request_post and 'fields_to_count' in request_post:
        for field in request_post['fields_to_count']:
            model = apps.get_model(field.split('.')[0], field.split('.')[1])
            query_set_name = filter.get_field_by_model(query_set.model, model)
            fields.append(
                {'verbose_name': model._meta.verbose_name,
                 'name': field.split('.')[1], 'prefix': '', 'field': 'count',
                 'model': model,
                 'query_set_name': query_set_name})
    # заголовок
    for col, field in enumerate(fields):
        ws.cell(1, col + 1).alignment = Alignment(wrap_text=True)
        ws.cell(1, col + 1).value = field['verbose_name']

    # данные
    for row, item in enumerate(query_set):
        for col, field in enumerate(fields):
            if field['field'] == 'custom':
                try:
                    val = ''
                    for c in item.children.all():
                        val = val + '; ' + c.doc_number
                    ws.cell(row + 2, col + 1).value = val
                except:
                    pass
            elif field['field'] == 'count':
                prefix = ''
                model_name = None
                field_model = field['model']
                temp_set = item.__getattribute__(field['query_set_name']).all()
                if getattr(item._meta.model, field['query_set_name']).field.model != field['model']:
                    field_model = getattr(item._meta.model, field['query_set_name']).field.model
                    model_name = field['name']
                    prefix = field['name'] + '__'
                    temp_set = item.__getattribute__(field['query_set_name']).filter(
                        doc_type=field['verbose_name'].lower())
                if request_post and 'count' in request_post and field['name'] in request_post['count']:
                    for field_key in request_post['count'][field['name']]:
                        temp_set = filter.add_filter(prefix + field_key, field_model, temp_set,
                                                     request_post['count'], model_name)
                ws.cell(row + 2, col + 1).value = str(temp_set.count())
            elif field['field'].__class__ == django.db.models.fields.related.ManyToManyField:
                # для полей типа ManyToMany просто перечисляем через запятую все найденные значения
                # это значит, что во второй модели должна быть прописана функция __str__
                val = ''
                f = item
                for p in str(field['prefix'] + field['name']).split('.'):
                    if f is None:
                        continue
                    f = f.__getattribute__(p)
                if f is None:
                    continue
                for i in f.all():
                    if val == '':
                        val = str(i)
                    else:
                        val = val + ',' + str(i)
                ws.cell(row + 2, col + 1).value = val
            elif field['field'].__class__ == django.db.models.fields.DateField:
                val = iggn_tools.tools.get_value(item, field['prefix'] + field['name'])
                if val:
                    val = datetime.strptime(iggn_tools.tools.get_value(item, field['prefix'] + field['name']), '%d.%m.%Y')
                    ws.cell(row + 2, col + 1).value = val
                    ws.cell(row + 2, col + 1).number_format = 'dd.mm.YYYY'
            else:
                # с этого момента начинаем брать каскадом значения полей
                # например field['prefix'] может быть равен 'organization.org_type.'
                # а field['name'] равен 'text'
                # в таком случае наша задача получить значение поля item.organization.org_type.text
                val = iggn_tools.tools.get_value(item, field['prefix'] + field['name'])
                ws.cell(row + 2, col + 1).value = val

        result.text = "Сделано %s из %s" % (row, count)
        result.save()
    result.text = 'Выгрузка таблицы "' + query_set.model._meta.verbose_name + '" с фильтрацией'
    result.datetime = timezone.now()
    result.file.save(uuid.uuid1().hex + '.xlsx', ContentFile(save_virtual_workbook(wb)))
    result.save()


def import_houses_from_licensing(file):
    file = 'C:\\1.xlsx'
    rb = xlrd.open_workbook(file)
    sheet = rb.sheet_by_name('Сведения об МКД')
    row = 8
    while row < sheet.nrows and sheet.row_values(row)[0] != '':
        row = row + 1
        val = sheet.row_values(row)
        area = val[1].strip()
        place = val[2].strip()
        city = val[3].strip()
        street = val[5].strip()
        number = val[6] if type(val[6]) is str else int(val[6])
        number = normalize_house_number(number)
        addr, created = dictionaries.models.Address.objects.get_or_create(area=area, place=place, city=city,
                                                                          street=street)
        house, created = dictionaries.models.House.objects.get_or_create(address=addr, number=number)
        try:
            org = dictionaries.models.Organization.objects.get(inn=int(val[16]))
            house.organization = org
        except dictionaries.models.Organization.DoesNotExist:
            pass
        house.building_year = int(val[7]) if val[7] else 0
        house.number_of_apartments = int(val[8]) if val[8] else 0
        house.total_area = val[9] if val[9] else 0
        house.living_area = val[10] if val[10] else 0
        house.non_living_area = val[11] if val[11] else 0
        house.changing_doc_number = val[19] if val[19] else None
        house.changing_doc_date = datetime(*xlrd.xldate_as_tuple(val[21], rb.datemode)) if val[21] else None
        house.changing_doc_header = val[22] if val[22] else None
        house.changing_org_date = datetime(*xlrd.xldate_as_tuple(val[24], rb.datemode)) if val[24] else None
        house.agr_conclusion_date = datetime(*xlrd.xldate_as_tuple(val[25], rb.datemode)) if val[25] else None
        house.management_start_date = datetime(*xlrd.xldate_as_tuple(val[26], rb.datemode)) if val[26] else None
        house.exclusion_date = datetime(*xlrd.xldate_as_tuple(val[28], rb.datemode)) if val[28] else None
        house.exclusion_legal_basis = val[29] if val[29] else None
        house.save()


def import_inspector_from_reestr_rasp(file):
    file = 'C:\\1.xlsx'
    rb = xlrd.open_workbook(file)
    sheet = rb.sheet_by_name('Лист1')
    row = 0
    while row < sheet.nrows and sheet.row_values(row)[0] != '':
        try:
            row = row + 1
            val = sheet.row_values(row)
            doc_number = str(val[1])
            doc_date = datetime(*xlrd.xldate_as_tuple(val[2], rb.datemode))
            name = val[6].strip()
            if name[-1] != '.':
                name = name + '.'
            print(f'{doc_number} {doc_date} {name}')
            try:
                insp = inspections.models.Inspection.objects.get(doc_number=doc_number, doc_date__year=doc_date.year)
            except inspections.models.Inspection.DoesNotExist:
                insp = inspections.models.Inspection(doc_number=doc_number, doc_date=doc_date)
                print('created')
            user = dictionaries.models.User.objects.get(shortname=name)
            insp.inspector = user
            insp.save()
            print(user.name)
        except (ValueError, TypeError):
            continue


def import_ad_from_penalty(file):
    file = 'C:\\1.xlsx'
    rb = xlrd.open_workbook(file)
    sheet = rb.sheet_by_name('Лист1')
    row = 0
    while row < sheet.nrows and sheet.row_values(row)[0] != '':
        try:
            row = row + 1
            val = sheet.row_values(row)
            doc_number = str(val[1])
            doc_date = datetime(*xlrd.xldate_as_tuple(val[2], rb.datemode))
            name = val[6].strip()
            if name[-1] != '.':
                name = name + '.'
            print(f'{doc_number} {doc_date} {name}')
            try:
                insp = inspections.models.Inspection.objects.get(doc_number=doc_number, doc_date__year=doc_date.year)
            except inspections.models.Inspection.DoesNotExist:
                insp = inspections.models.Inspection(doc_number=doc_number, doc_date=doc_date)
                print('created')
            user = dictionaries.models.User.objects.get(shortname=name)
            insp.inspector = user
            insp.save()
            print(user.name)
        except (ValueError, TypeError):
            continue


def import_from_reestr_tsj():
    rb = xlrd.open_workbook('C:/Users/ivsemionov/Desktop/tsj.xlsx')
    sheet = rb.sheet_by_index(0)
    row = 3
    while row < sheet.nrows and sheet.row_values(row)[11] != '':
        print(str(row) + " " + sheet.row_values(row)[11])
        try:
            org, created = dictionaries.models.Organization.objects.get_or_create(inn=int(sheet.row_values(row)[15]))
            org.name = sheet.row_values(row)[11]
            org.ogrn = int(sheet.row_values(row)[14])
            org.save()
        except: pass
        row = row + 1
