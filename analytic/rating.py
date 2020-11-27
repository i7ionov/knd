"""НЕ АКТУАЛЬНО!!!!"""
import openpyxl  # openpyxl==2.5.14
import datetime
from inspections.models import Inspection, Precept

date = datetime.date.today()
wb = openpyxl.load_workbook('C:\\Users\\ivsemionov\\Desktop\\В работе\\Данные для рейтинга.xlsx')
sheet = wb.worksheets[0]
row = 6
while sheet.cell(row, 4).value:
    precept_row1 = 13
    precept_row2 = 14
    inn = sheet.cell(row, 4).value
    sheet.cell(row, 7).value = Inspection.objects.filter(organization__inn=inn,
                                                         act_date__range=('2019-10-01', '2020-01-01'),
                                                         inspection_result__id=1,
                                                         inspection_type__id=1).count()
    sheet.cell(row, 10).value = Precept.objects.filter(organization__inn=inn,
                                                         doc_date__range=('2019-10-01', '2020-01-01')).count()
    sheet.cell(row, 13).value = Precept.objects.filter(organization__inn=inn,
                                                       children__doc_type='проверка',
                                                       children__inspection__inspection_result__id=1,
                                                       doc_date__range=('2019-10-01', '2020-01-01'),
                                                       precept_result__id=1).count()
    sheet.cell(row, 14).value = Precept.objects.filter(organization__inn=inn,
                                                       children__doc_type='проверка',
                                                       children__inspection__inspection_result__id=1,
                                                       doc_date__range=('2019-10-01', '2020-01-01')).count()
    row += 1
wb.save('C:\\Users\\ivsemionov\\Desktop\\В работе\\1.xlsx')